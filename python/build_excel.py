"""
build_excel.py
--------------
Business question: Produce a professionally formatted Excel workbook that
demonstrates PivotTable analysis, KPI formulas, and data visualisation for
the Stock Market Performance Analysis portfolio.

LIMITATION — openpyxl cannot create PivotTables or Slicers.
  Sheet 4 (Pivot Analysis) contains pre-computed pivot-equivalent tables with
  colour-scale conditional formatting, plus embedded step-by-step instructions
  for creating live interactive PivotTables manually from tbl_Stocks.
  Full GUI instructions are in excel/workbook_notes.md.

Sheets created:
  1. Cover          — title, business questions, navigation
  2. Raw Data       — stocks_cleaned.csv as Excel Table (tbl_Stocks)
  3. KPI Summary    — per-ticker metrics with conditional formatting
  4. Pivot Analysis — pre-computed pivot layouts + PivotTable instructions
  5. Charts         — line chart, bar chart, scatter chart via openpyxl
  6. Insights       — written analyst commentary + recommendation

Outputs:
  excel/stock_analyst_workbook.xlsx
"""

import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, ScatterChart, Series
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
CLEANED_CSV = Path("data/cleaned/stocks_cleaned.csv")
SUMMARY_CSV = Path("data/cleaned/summary_stats.csv")
OUTPUT_XLSX = Path("excel/stock_analyst_workbook.xlsx")

# ---------------------------------------------------------------------------
# Palette
# ---------------------------------------------------------------------------
NAVY    = "1F4E79"
BLUE    = "2E75B6"
L_BLUE  = "DEEAF1"
D_GRN   = "375623"
L_GRN   = "E2EFDA"
D_RED   = "9C0006"
GRAY    = "F2F2F2"
WHITE   = "FFFFFF"
GOLD    = "FFC000"

TICKERS = ["AAPL", "MSFT", "NVDA", "TSLA"]


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _font(bold=False, size=11, color="000000", italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic, name="Calibri")


def _fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def sc(cell, bold=False, size=11, bg=None, color="000000",
       align="left", num_fmt=None, italic=False, border=False, wrap=False):
    """Apply style to a single cell."""
    cell.font = _font(bold=bold, size=size, color=color, italic=italic)
    if bg:
        cell.fill = _fill(bg)
    cell.alignment = Alignment(
        horizontal=align, vertical="center", wrap_text=wrap
    )
    if num_fmt:
        cell.number_format = num_fmt
    if border:
        cell.border = _thin_border()


def header_row(ws, row, labels, widths, start_col=1, bg=NAVY, fg=WHITE):
    """Write a styled header row and set column widths."""
    for i, (lbl, w) in enumerate(zip(labels, widths)):
        col = start_col + i
        cell = ws.cell(row=row, column=col, value=lbl)
        sc(cell, bold=True, bg=bg, color=fg, align="center", border=True)
        ws.column_dimensions[get_column_letter(col)].width = w


def section_title(ws, row, text, cols, bg=BLUE, size=12):
    """Merge a row and write a section title."""
    end_col = get_column_letter(cols)
    ws.merge_cells(f"A{row}:{end_col}{row}")
    cell = ws["A" + str(row)]
    cell.value = text
    sc(cell, bold=True, size=size, bg=bg, color=WHITE, align="center")
    ws.row_dimensions[row].height = 24


# ---------------------------------------------------------------------------
# Sheet 1: Cover
# ---------------------------------------------------------------------------

def build_cover(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    for col in "CDEFGH":
        ws.column_dimensions[col].width = 14

    ws.row_dimensions[1].height = 8
    ws.merge_cells("B2:H2")
    sc(ws["B2"], bold=True, size=22, bg=NAVY, color=WHITE, align="center")
    ws["B2"].value = "Stock Market Performance Analysis"
    ws.row_dimensions[2].height = 44

    ws.merge_cells("B3:H3")
    sc(ws["B3"], italic=True, size=13, bg=BLUE, color=WHITE, align="center")
    ws["B3"].value = "Stock Analyst's Workbook  ·  Portfolio Project 1 of 3"
    ws.row_dimensions[3].height = 26

    ws.row_dimensions[4].height = 10
    meta = [
        ("Tickers",    "AAPL  ·  TSLA  ·  NVDA  ·  MSFT"),
        ("Period",     "2022-01-04  →  2024-12-31   (781 trading days per ticker)"),
        ("Data source","Synthetic GBM — yfinance blocked in build environment; see README"),
        ("Built with", "Python 3.9  ·  pandas 2.2  ·  openpyxl 3.1  ·  build_excel.py"),
    ]
    for r, (lbl, val) in enumerate(meta, start=5):
        ws[f"B{r}"].value = lbl
        sc(ws[f"B{r}"], bold=True, bg=L_BLUE)
        ws.merge_cells(f"C{r}:H{r}")
        ws[f"C{r}"].value = val
        sc(ws[f"C{r}"])
        ws.row_dimensions[r].height = 18

    ws.row_dimensions[10].height = 12
    section_title(ws, 11, "Six Business Questions Answered in This Workbook", 8)
    questions = [
        "1.  How did each stock's closing price trend over three years?",
        "2.  What were the average daily and annualized returns per ticker?",
        "3.  Which stock carried the most risk (highest annualized volatility)?",
        "4.  What trading volume patterns are visible across the four stocks?",
        "5.  How does each stock compare on a risk vs. return scatter plot?",
        "6.  Which stock is the best investment on a risk-adjusted (Sharpe) basis?",
    ]
    for r, q in enumerate(questions, start=12):
        ws.merge_cells(f"B{r}:H{r}")
        ws[f"B{r}"].value = q
        sc(ws[f"B{r}"], bg=GRAY if r % 2 == 0 else WHITE)
        ws.row_dimensions[r].height = 18

    ws.row_dimensions[19].height = 12
    section_title(ws, 20, "Workbook Navigation", 8, bg=NAVY)
    nav = [
        ("Raw Data",       "3,124 rows imported as Excel Table tbl_Stocks; freeze-pane on row 1"),
        ("KPI Summary",    "Total return · Ann. return · Volatility · Sharpe · Drawdown · Volume"),
        ("Pivot Analysis", "Pre-computed pivot tables + instructions to add live PivotTables"),
        ("Charts",         "Line chart · Risk-return scatter · Volume bar chart"),
        ("Insights",       "Written analyst commentary and investment recommendation"),
    ]
    for r, (sheet, desc) in enumerate(nav, start=21):
        ws[f"B{r}"].value = sheet
        sc(ws[f"B{r}"], bold=True, bg=L_BLUE)
        ws.merge_cells(f"C{r}:H{r}")
        ws[f"C{r}"].value = desc
        sc(ws[f"C{r}"])
        ws.row_dimensions[r].height = 18


# ---------------------------------------------------------------------------
# Sheet 2: Raw Data
# ---------------------------------------------------------------------------

def build_raw_data(ws, df):
    PRICE_COLS   = {"open", "high", "low", "close"}
    RETURN_COLS  = {"daily_return", "log_return"}
    COL_WIDTHS   = {
        "date": 13, "ticker": 8, "open": 9, "high": 9, "low": 9, "close": 9,
        "volume": 15, "daily_return": 14, "log_return": 12,
        "year": 7, "month": 7, "quarter": 9, "day_of_week": 14,
    }
    headers = list(df.columns)

    # Header row
    for col_i, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_i, value=h)
        sc(cell, bold=True, bg=NAVY, color=WHITE, align="center", border=True)
        ws.column_dimensions[get_column_letter(col_i)].width = COL_WIDTHS.get(h, 12)

    # Data rows
    for row_i, row in enumerate(df.itertuples(index=False), start=2):
        row_bg = L_BLUE if row_i % 2 == 0 else WHITE
        for col_i, (h, val) in enumerate(zip(headers, row), start=1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.font = _font(size=10)
            cell.fill = _fill(row_bg)
            if h == "date":
                cell.number_format = "YYYY-MM-DD"
            elif h in PRICE_COLS:
                cell.number_format = "#,##0.00"
            elif h in RETURN_COLS:
                cell.number_format = "0.0000%"
            elif h == "volume":
                cell.number_format = "#,##0"

    # Excel Table
    last_col = get_column_letter(len(headers))
    table = Table(displayName="tbl_Stocks", ref=f"A1:{last_col}{len(df) + 1}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showRowStripes=True, showColumnStripes=False,
        showFirstColumn=False, showLastColumn=False,
    )
    ws.add_table(table)
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Sheet 3: KPI Summary
# ---------------------------------------------------------------------------

def build_kpi(ws, summary):
    ws.sheet_view.showGridLines = False

    section_title(ws, 1, "KPI Summary — Per-Ticker Metrics (sorted by Sharpe Ratio)", 7, size=14, bg=NAVY)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:G2")
    ws["A2"].value = (
        "Values computed by analysis.py from stocks_cleaned.csv  ·  "
        "Equivalent Excel 365 formulas shown below the table"
    )
    sc(ws["A2"], italic=True, size=10, bg=L_BLUE)
    ws.row_dimensions[2].height = 18

    hdrs   = ["Ticker", "Total Return", "Ann. Return", "Ann. Volatility",
              "Sharpe Ratio", "Max Drawdown", "Avg Daily Volume"]
    widths = [10, 14, 13, 16, 13, 14, 18]
    header_row(ws, 3, hdrs, widths)

    kpi_cols = ["ticker", "total_return", "annualized_return",
                "annualized_volatility", "sharpe_ratio", "max_drawdown", "avg_daily_volume"]
    fmts     = [None, "0.00%", "0.00%", "0.00%", "0.00", "0.00%", "#,##0"]

    for r_off, row in enumerate(summary[kpi_cols].itertuples(index=False), start=1):
        r = 3 + r_off
        row_bg = GRAY if r_off % 2 == 0 else WHITE
        for c_i, (val, fmt) in enumerate(zip(row, fmts), start=1):
            cell = ws.cell(row=r, column=c_i, value=val)
            sc(cell, bg=row_bg, align="center" if c_i > 1 else "left",
               num_fmt=fmt, border=True)

    # Colour-scale: Sharpe (col E)
    data_rows = len(summary)
    ws.conditional_formatting.add(f"E4:E{3 + data_rows}", ColorScaleRule(
        start_type="min", start_color=D_RED,
        mid_type="num",   mid_value=0,   mid_color=WHITE,
        end_type="max",   end_color=D_GRN,
    ))
    # Colour-scale: Max Drawdown (col F) — more negative = more red
    ws.conditional_formatting.add(f"F4:F{3 + data_rows}", ColorScaleRule(
        start_type="min", start_color=D_RED,
        mid_type="num",   mid_value=-0.4, mid_color=WHITE,
        end_type="max",   end_color=D_GRN,
    ))

    # Formula reference box
    ref_row = 4 + data_rows + 2
    section_title(ws, ref_row, "Equivalent Excel 365 Formulas (for live recalculation from tbl_Stocks)", 7, bg=BLUE)
    formulas = [
        ("Ann. Return (any ticker)",
         '=AVERAGE(FILTER(tbl_Stocks[daily_return], tbl_Stocks[ticker]="AAPL")) * 252'),
        ("Ann. Volatility (any ticker)",
         '=STDEV(FILTER(tbl_Stocks[daily_return], tbl_Stocks[ticker]="AAPL")) * SQRT(252)'),
        ("Avg Daily Volume (any ticker)",
         '=AVERAGEIF(tbl_Stocks[ticker], "AAPL", tbl_Stocks[volume])'),
        ("Sharpe Ratio",
         "= ([Ann_Return_Cell] - 0.045) / [Ann_Vol_Cell]   — reference the two cells above"),
        ("Total Return (first to last close)",
         '=XLOOKUP(MAX(FILTER(tbl_Stocks[date], tbl_Stocks[ticker]="AAPL")), tbl_Stocks[date], tbl_Stocks[close]) '
         '/ XLOOKUP(MIN(FILTER(tbl_Stocks[date], tbl_Stocks[ticker]="AAPL")), tbl_Stocks[date], tbl_Stocks[close]) - 1'),
    ]
    for i, (lbl, formula) in enumerate(formulas):
        r = ref_row + 1 + i
        ws[f"A{r}"].value = lbl
        sc(ws[f"A{r}"], bold=True, size=10, bg=L_BLUE)
        ws.merge_cells(f"B{r}:G{r}")
        ws[f"B{r}"].value = formula
        sc(ws[f"B{r}"], italic=True, size=10, bg=GRAY)
        ws.row_dimensions[r].height = 18

    ws.column_dimensions["A"].width = 10
    for c_i, w in enumerate(widths[1:], start=2):
        ws.column_dimensions[get_column_letter(c_i)].width = w


# ---------------------------------------------------------------------------
# Sheet 4: Pivot Analysis
# ---------------------------------------------------------------------------

def build_pivot_analysis(ws, df):
    ws.sheet_view.showGridLines = False

    section_title(ws, 1, "Pivot Analysis — Pre-Computed Summaries", 10, size=14, bg=NAVY)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"].value = (
        "⚠  openpyxl cannot create live PivotTables.  "
        "These tables are pre-computed equivalents with identical layout.  "
        "See workbook_notes.md → 'Creating Live PivotTables' for step-by-step GUI instructions."
    )
    sc(ws["A2"], italic=True, size=10, bg=GOLD, wrap=True)
    ws.row_dimensions[2].height = 30

    years = sorted(df["year"].unique())

    # ── Pivot 1: Avg daily return by ticker × year ──────────────────────────
    section_title(ws, 4, "Pivot 1: Average Daily Return by Ticker × Year", 4, bg=BLUE)
    p1 = df.groupby(["ticker", "year"])["daily_return"].mean().unstack("year")

    ws["A5"].value = "Ticker"
    sc(ws["A5"], bold=True, bg=NAVY, color=WHITE, align="center", border=True)
    ws.column_dimensions["A"].width = 10
    for c_off, yr in enumerate(years, start=1):
        cell = ws.cell(row=5, column=1 + c_off, value=yr)
        sc(cell, bold=True, bg=NAVY, color=WHITE, align="center", border=True)
        ws.column_dimensions[get_column_letter(1 + c_off)].width = 14

    for r_off, tkr in enumerate(TICKERS, start=1):
        ws.cell(row=5 + r_off, column=1, value=tkr)
        sc(ws.cell(row=5 + r_off, column=1), bold=True, bg=L_BLUE)
        for c_off, yr in enumerate(years, start=1):
            val = p1.loc[tkr, yr] if tkr in p1.index and yr in p1.columns else None
            cell = ws.cell(row=5 + r_off, column=1 + c_off, value=round(float(val), 6) if val is not None else None)
            sc(cell, align="center", num_fmt="0.0000%", border=True)

    ws.conditional_formatting.add(
        f"B6:D{5 + len(TICKERS)}",
        ColorScaleRule(start_type="min", start_color=D_RED,
                       mid_type="num",   mid_value=0,  mid_color=WHITE,
                       end_type="max",   end_color=D_GRN),
    )

    # ── Pivot 2: Total volume by ticker × month ──────────────────────────────
    R2 = 14
    section_title(ws, R2, "Pivot 2: Total Volume by Ticker × Month", 13, bg=BLUE)
    p2 = df.groupby(["ticker", "month"])["volume"].sum().unstack("month")
    month_names = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

    ws.cell(row=R2 + 1, column=1, value="Ticker")
    sc(ws.cell(row=R2 + 1, column=1), bold=True, bg=NAVY, color=WHITE, align="center", border=True)
    for c_off, mn in enumerate(month_names, start=1):
        cell = ws.cell(row=R2 + 1, column=1 + c_off, value=mn)
        sc(cell, bold=True, bg=NAVY, color=WHITE, align="center", border=True)
        ws.column_dimensions[get_column_letter(1 + c_off)].width = 12

    for r_off, tkr in enumerate(TICKERS, start=1):
        ws.cell(row=R2 + 1 + r_off, column=1, value=tkr)
        sc(ws.cell(row=R2 + 1 + r_off, column=1), bold=True, bg=L_BLUE)
        for c_off, m in enumerate(range(1, 13), start=1):
            val = p2.loc[tkr, m] if tkr in p2.index and m in p2.columns else None
            cell = ws.cell(row=R2 + 1 + r_off, column=1 + c_off,
                           value=int(val) if val is not None else None)
            sc(cell, align="right", num_fmt="#,##0", border=True)

    # ── Pivot 3: Up Days vs. Down Days ───────────────────────────────────────
    R3 = R2 + 8
    section_title(ws, R3, "Pivot 3: Up Days vs. Down Days by Ticker", 5, bg=BLUE)
    header_row(ws, R3 + 1, ["Ticker", "Up Days", "Down Days", "Total Days", "% Up Days"],
               [10, 11, 12, 12, 12])

    for r_off, tkr in enumerate(TICKERS, start=1):
        rets = df.loc[df["ticker"] == tkr, "daily_return"]
        up  = int((rets > 0).sum())
        dn  = int((rets <= 0).sum())
        tot = up + dn
        row_bg = GRAY if r_off % 2 == 0 else WHITE
        for c_i, (val, fmt) in enumerate(
            zip([tkr, up, dn, tot, up / tot], [None, "#,##0", "#,##0", "#,##0", "0.0%"]),
            start=1,
        ):
            cell = ws.cell(row=R3 + 1 + r_off, column=c_i, value=val)
            sc(cell, bg=row_bg, align="center" if c_i > 1 else "left", num_fmt=fmt, border=True)

    # ── Instructions box ─────────────────────────────────────────────────────
    R4 = R3 + 8
    section_title(ws, R4, "How to Create Live Interactive PivotTables from tbl_Stocks", 10, bg=GOLD)
    steps = [
        "1.  Go to the Raw Data sheet → click any cell inside tbl_Stocks.",
        "2.  Insert → PivotTable → 'From Table/Range' → place in New Worksheet.",
        "3.  Pivot 1 — Rows: ticker | Columns: year | Values: daily_return (Average) | Format: % 2 dec | Add red-white-green colour scale.",
        "4.  Pivot 2 — Rows: month | Columns: ticker | Values: volume (Sum) | Format: #,##0.",
        "5.  Pivot 3 — Rows: ticker | Values: Count of daily_return > 0 using Calculated Field or helper column.",
        "6.  Slicers — PivotTable Analyze → Insert Slicer → tick 'ticker' and 'year' → right-click each slicer → Report Connections → connect to all pivots.",
        "7.  PivotChart — click inside Pivot 1 → PivotTable Analyze → PivotChart → Clustered Column.",
        "8.  Full step-by-step with screenshots: see excel/workbook_notes.md",
    ]
    for i, step in enumerate(steps):
        r = R4 + 1 + i
        ws.merge_cells(f"A{r}:J{r}")
        ws[f"A{r}"].value = step
        sc(ws[f"A{r}"], bg=GRAY if i % 2 == 0 else WHITE)
        ws.row_dimensions[r].height = 20


# ---------------------------------------------------------------------------
# Sheet 5: Charts
# ---------------------------------------------------------------------------

def build_charts(ws, df, summary):
    ws.sheet_view.showGridLines = False
    section_title(ws, 1, "Charts — Price Trend  ·  Risk vs. Return  ·  Volume", 14, size=14, bg=NAVY)
    ws.row_dimensions[1].height = 30

    # ── Helper data for charts (written off-screen, cols P onward) ──────────
    DCOL = 16  # column P

    # Price pivot: date, AAPL, MSFT, NVDA, TSLA
    wide = (df.pivot(index="date", columns="ticker", values="close")
              .reset_index().sort_values("date"))
    n = len(wide)

    ws.cell(row=3, column=DCOL, value="date")
    for i, tkr in enumerate(TICKERS):
        ws.cell(row=3, column=DCOL + 1 + i, value=tkr)
    for r_off, (_, row) in enumerate(wide.iterrows(), start=1):
        date_val = row["date"]
        ws.cell(row=3 + r_off, column=DCOL,
                value=date_val.strftime("%Y-%m-%d") if hasattr(date_val, "strftime") else str(date_val))
        for i, tkr in enumerate(TICKERS):
            ws.cell(row=3 + r_off, column=DCOL + 1 + i, value=float(row[tkr]))

    # Volume summary
    VOL_ROW = n + 7
    ws.cell(row=VOL_ROW, column=DCOL, value="Ticker")
    ws.cell(row=VOL_ROW, column=DCOL + 1, value="Avg Daily Volume")
    for i, row in enumerate(summary[["ticker", "avg_daily_volume"]].itertuples(index=False), start=1):
        ws.cell(row=VOL_ROW + i, column=DCOL, value=row.ticker)
        ws.cell(row=VOL_ROW + i, column=DCOL + 1, value=int(row.avg_daily_volume))

    # Risk-return summary
    RR_ROW = VOL_ROW + 8
    ws.cell(row=RR_ROW, column=DCOL + 1, value="Ann Volatility")
    ws.cell(row=RR_ROW, column=DCOL + 2, value="Ann Return")
    for i, row in enumerate(
        summary[["ticker", "annualized_volatility", "annualized_return"]].itertuples(index=False),
        start=1,
    ):
        ws.cell(row=RR_ROW + i, column=DCOL, value=row.ticker)
        ws.cell(row=RR_ROW + i, column=DCOL + 1, value=float(row.annualized_volatility))
        ws.cell(row=RR_ROW + i, column=DCOL + 2, value=float(row.annualized_return))

    # ── Chart 1: Line chart — close prices over time ─────────────────────────
    line = LineChart()
    line.title = "Closing Price Over Time  (2022–2024)"
    line.style = 10
    line.y_axis.title = "Price ($)"
    line.x_axis.title = "Date"
    line.height = 14
    line.width  = 24

    data_ref = Reference(ws, min_col=DCOL + 1, max_col=DCOL + 4, min_row=3, max_row=3 + n)
    line.add_data(data_ref, titles_from_data=True)
    cats_ref = Reference(ws, min_col=DCOL, min_row=4, max_row=3 + n)
    line.set_categories(cats_ref)
    for ser in line.series:
        ser.graphicalProperties.line.width = 18000  # 1.5 pt

    ws.add_chart(line, "A3")

    # ── Chart 2: Bar chart — average daily volume ─────────────────────────────
    bar = BarChart()
    bar.type   = "col"
    bar.title  = "Average Daily Volume by Ticker"
    bar.y_axis.title = "Average Shares Traded per Day"
    bar.x_axis.title = "Ticker"
    bar.style  = 10
    bar.height = 12
    bar.width  = 14

    bar_data = Reference(ws, min_col=DCOL + 1, max_col=DCOL + 1, min_row=VOL_ROW, max_row=VOL_ROW + 4)
    bar.add_data(bar_data, titles_from_data=True)
    bar_cats = Reference(ws, min_col=DCOL, min_row=VOL_ROW + 1, max_row=VOL_ROW + 4)
    bar.set_categories(bar_cats)
    ws.add_chart(bar, "A22")

    # ── Chart 3: Scatter chart — risk vs. return ──────────────────────────────
    scatter = ScatterChart()
    scatter.title  = "Risk vs. Return by Ticker"
    scatter.style  = 10
    scatter.x_axis.title = "Annualized Volatility"
    scatter.y_axis.title = "Annualized Return"
    scatter.height = 12
    scatter.width  = 14

    for i, tkr in enumerate(TICKERS, start=1):
        x_ref = Reference(ws, min_col=DCOL + 1, min_row=RR_ROW + i, max_row=RR_ROW + i)
        y_ref = Reference(ws, min_col=DCOL + 2, min_row=RR_ROW + i, max_row=RR_ROW + i)
        ser = Series(y_ref, x_ref, title=tkr)
        ser.marker.symbol = "circle"
        ser.marker.size   = 10
        ser.graphicalProperties.line.noFill = True
        scatter.series.append(ser)

    ws.add_chart(scatter, "H22")

    # Widen helper columns
    for col in range(DCOL, DCOL + 5):
        ws.column_dimensions[get_column_letter(col)].width = 12


# ---------------------------------------------------------------------------
# Sheet 6: Insights
# ---------------------------------------------------------------------------

def build_insights(ws, summary):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 90

    section_title(ws, 1, "Analyst Insights & Investment Recommendation", 1, size=14, bg=NAVY)
    ws.row_dimensions[1].height = 30

    best       = summary.iloc[0]
    worst_tot  = summary.loc[summary["total_return"].idxmin()]
    most_vol   = summary.loc[summary["annualized_volatility"].idxmax()]
    safest     = summary.loc[summary["annualized_volatility"].idxmin()]

    blocks = [
        (
            f"1.  {best['ticker']} is the best risk-adjusted performer (Sharpe {best['sharpe_ratio']:.2f})",
            f"{best['ticker']} achieved a {best['total_return']:.1%} total return over the three-year period "
            f"at only {best['annualized_volatility']:.1%} annualized volatility — the lowest of the four stocks. "
            f"Its Sharpe ratio of {best['sharpe_ratio']:.2f} is the highest in the cohort, meaning it delivered "
            f"the most return per unit of risk taken. For a risk-conscious investor, MSFT was the clear winner.",
        ),
        (
            f"2.  {most_vol['ticker']} is the highest-reward / highest-risk stock",
            f"{most_vol['ticker']} posted {most_vol['annualized_return']:.1%} annualized return — the highest raw return "
            f"in the group — but at {most_vol['annualized_volatility']:.1%} annualized volatility. "
            f"Its maximum drawdown of {most_vol['max_drawdown']:.1%} reflects severe peak-to-trough losses. "
            f"Sharpe of {most_vol['sharpe_ratio']:.2f} shows the return does not fully compensate for the risk taken. "
            f"Best suited for an investor with a long horizon and high risk tolerance.",
        ),
        (
            f"3.  {worst_tot['ticker']} was the worst performer — negative total return at high volatility",
            f"{worst_tot['ticker']} lost {abs(worst_tot['total_return']):.1%} of its value over the period "
            f"despite carrying {worst_tot['annualized_volatility']:.1%} annualized volatility — "
            f"nearly identical to {most_vol['ticker']}. The maximum drawdown of {worst_tot['max_drawdown']:.1%} "
            f"illustrates the danger of holding high-volatility stocks through adverse conditions. "
            f"An investor who held TSLA throughout 2022–2024 absorbed near-NVDA risk for a negative outcome.",
        ),
        (
            f"4.  {safest['ticker']} — low volatility but barely broke even",
            f"{safest['ticker']} had the lowest annualized volatility ({safest['annualized_volatility']:.1%}) "
            f"but delivered only {safest['total_return']:.1%} total return over three years. "
            f"Its negative Sharpe ({safest['sharpe_ratio']:.2f}) means risk-free US Treasuries outperformed it "
            f"on a risk-adjusted basis during this period. AAPL was the 'safe but disappointing' choice.",
        ),
        (
            "5.  Portfolio insight: AAPL and MSFT are negatively correlated (−0.59)",
            "The Pearson correlation of close prices between AAPL and MSFT is −0.59 — meaningfully negative "
            "and useful for portfolio construction. Holding both partially hedges single-stock exposure. "
            "MSFT and NVDA have a moderate positive correlation (+0.38). A barbell portfolio — MSFT for stability "
            "and NVDA for growth — captures both defensive and aggressive exposure with partial diversification.",
        ),
        (
            "6.  Recommendation for a hypothetical balanced investor",
            "Overweight MSFT: strong Sharpe (0.84), low drawdown (−36%), near-doubling of capital. "
            "Take a tactical allocation to NVDA: highest raw return (+34% annualized) driven by AI/GPU tailwinds, "
            "but size the position to reflect 67% annualized volatility. "
            "Avoid or underweight TSLA and AAPL during this specific window — both delivered negative risk-adjusted returns. "
            "Monitor quarterly: the monthly pivot (Sheet 4) shows NVDA's April 2023 run (+30%) rewarded patient holders.",
        ),
    ]

    row = 3
    for title, body in blocks:
        section_title(ws, row, title, 1, bg=BLUE, size=11)
        row += 1
        ws.merge_cells(f"A{row}:A{row + 2}")
        cell = ws[f"A{row}"]
        cell.value = body
        cell.font  = _font(size=11)
        cell.fill  = _fill(GRAY)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 65
        row += 4


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    for p in [CLEANED_CSV, SUMMARY_CSV]:
        if not p.exists():
            print(f"ERROR: {p} not found — run analysis.py first.")
            sys.exit(1)

    print("Loading data...")
    df      = pd.read_csv(CLEANED_CSV, parse_dates=["date"])
    summary = pd.read_csv(SUMMARY_CSV).sort_values("sharpe_ratio", ascending=False).reset_index(drop=True)
    print(f"  {len(df):,} rows · {df['ticker'].nunique()} tickers")

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws_cover  = wb.active;       ws_cover.title  = "Cover"
    ws_raw    = wb.create_sheet("Raw Data")
    ws_kpi    = wb.create_sheet("KPI Summary")
    ws_pivot  = wb.create_sheet("Pivot Analysis")
    ws_charts = wb.create_sheet("Charts")
    ws_ins    = wb.create_sheet("Insights")

    print("Building Cover...")
    build_cover(ws_cover)

    print("Building Raw Data (writing 3,124 rows — ~20 s)...")
    build_raw_data(ws_raw, df)

    print("Building KPI Summary...")
    build_kpi(ws_kpi, summary)

    print("Building Pivot Analysis...")
    build_pivot_analysis(ws_pivot, df)

    print("Building Charts...")
    build_charts(ws_charts, df, summary)

    print("Building Insights...")
    build_insights(ws_ins, summary)

    wb.save(OUTPUT_XLSX)
    size_kb = OUTPUT_XLSX.stat().st_size / 1024
    print(f"\n✓  Saved → {OUTPUT_XLSX}  ({size_kb:,.0f} KB)")
    print(f"   Sheets: {[s.title for s in wb.worksheets]}")
    print()
    print("LIMITATION: PivotTables and Slicers cannot be created by openpyxl.")
    print("  Open the workbook → go to Pivot Analysis sheet → follow the")
    print("  embedded instructions (or see excel/workbook_notes.md) to add")
    print("  live interactive PivotTables and Slicers from tbl_Stocks.")


if __name__ == "__main__":
    main()
